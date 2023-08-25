from __future__ import print_function
import argparse
import pandas as pd
import os
import os.path as osp
import copy
import time
import random
import numpy as np

import torch
import torch.nn as nn
import torch.optim as optim
from torchvision import transforms
from torch.utils.tensorboard import SummaryWriter
from torch.utils.data import DataLoader

from dataloader import LLP_dataset, ToTensor, categories
from nets.net_audiovisual import NREP, LabelSmoothingNCELoss
from utils.eval_metrics import segment_level, event_level, print_overall_metric

from utils.logger import get_logger
from openpyxl import load_workbook
from nets.criterion import EvidenceLoss, MutualLearningLoss
from tqdm import tqdm
import datetime
import pickle as pkl
import pdb

exp_logger = get_logger()



def get_LLP_dataloader(args):
    train_dataset = LLP_dataset(label=args.label_train, audio_dir=args.audio_dir,
                                video_dir=args.video_dir, st_dir=args.st_dir,
                                transform=transforms.Compose([ToTensor()]),
                                v_pseudo_data_dir=args.v_pseudo_data_dir, a_pseudo_data_dir=args.a_pseudo_data_dir)
    val_dataset = LLP_dataset(label=args.label_val, audio_dir=args.audio_dir,
                              video_dir=args.video_dir, st_dir=args.st_dir,
                              transform=transforms.Compose([ToTensor()]), v_pseudo_data_dir=args.v_pseudo_data_dir, a_pseudo_data_dir=args.a_pseudo_data_dir)
    test_dataset = LLP_dataset(label=args.label_test, audio_dir=args.audio_dir,
                               video_dir=args.video_dir, st_dir=args.st_dir,
                               transform=transforms.Compose([ToTensor()]), v_pseudo_data_dir=args.v_pseudo_data_dir, a_pseudo_data_dir=args.a_pseudo_data_dir)

    train_loader = DataLoader(train_dataset, batch_size=args.batch_size,
                              shuffle=True, num_workers=5, pin_memory=True, sampler=None)
    val_loader = DataLoader(val_dataset, batch_size=1, shuffle=False,
                            num_workers=1, pin_memory=True, sampler=None)
    test_loader = DataLoader(test_dataset, batch_size=1, shuffle=False,
                             num_workers=1, pin_memory=True, sampler=None)

    return train_loader, val_loader, test_loader


def set_random_seed(seed):
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)
    random.seed(seed)
    np.random.seed(seed)


def get_random_state():
    state = {
        'torch_rng': torch.get_rng_state(),
        'cuda_rng': torch.cuda.get_rng_state(),
        'random_rng': random.getstate(),
        'numpy_rng': np.random.get_state()
    }
    return state



def train_with_pseudo_labels(args, models, train_loader, optimizers, criterion, epoch, logger):
    exp_logger(f"begin train_with_pseudo_labels.")
    for model in models:
        model.train()

    criterion2 = LabelSmoothingNCELoss(classes=10, smoothing=args.nce_smooth)
    criterion3 = MutualLearningLoss()
    criterion_edl = EvidenceLoss(num_classes=2, evidence = args.evidence, loss_type= args.loss_type, uncertain=args.uncertain)

    for batch_idx, sample in tqdm(enumerate(train_loader), total=len(train_loader), desc="Training with pseudo labels epoch {}".format(epoch)):
        audio, video, video_st, target = sample['audio'].to('cuda'), \
                                         sample['video_s'].to('cuda'), \
                                         sample['video_st'].to('cuda'), \
                                         sample['label'].type(torch.FloatTensor).to('cuda')
        Pa, Pv = sample['Pa'].type(torch.FloatTensor).to('cuda'), sample['Pv'].type(torch.FloatTensor).to('cuda')

        audio_pseudo_labels = sample['audio_pseudo_labels'].float().to('cuda')     # (B, 10, 25)
        visual_pseudo_labels = sample['visual_pseudo_labels'].float().to('cuda')    # (B, 10, 25)

        for optimizer in optimizers:
            optimizer.zero_grad()

        if epoch <= args.epochs:
            ## Train audio model
            global_logits, a_logits, v_logits, frame_logits, temporal_logits, frame_att, temporal_att, sims_after, mask_after = models[0](audio, video, video_st, with_ca=True, epoch_num = epoch-1)

            a_frame_logits = frame_logits[:,:,0,:,:]
            v_frame_logits = frame_logits[:,:,1,:,:]
            fbcl_loss = args.lamda_1*criterion2(sims_after, mask_after) + args.lamda_2*(criterion3(temporal_att[..., 0], frame_att[..., 0].max(dim=-2)[0]) + criterion3(temporal_att[..., 1], frame_att[..., 1].max(dim=-2)[0]))

            frame_target = audio_pseudo_labels * visual_pseudo_labels * target.unsqueeze(dim=1).expand(Pa.shape[0], 10, 25)
            global_sl_loss = criterion_edl(global_logits, target)
            temporal_sl_loss = criterion_edl(temporal_logits, frame_target) 
            tel_loss = global_sl_loss + temporal_sl_loss

            a_frame_sl_loss = criterion_edl(a_frame_logits, audio_pseudo_labels, fore_only = True)
            v_frame_sl_loss = criterion_edl(v_frame_logits, visual_pseudo_labels, fore_only = True)
            ps_loss = args.audio_factor * a_frame_sl_loss + v_frame_sl_loss

            audio_sl_loss = criterion_edl(a_logits, Pa)
            video_sl_loss = criterion_edl(v_logits, Pv)
            mel_loss = args.audio_factor * audio_sl_loss + video_sl_loss

            loss = ps_loss + args.lamda_3*mel_loss + args.lamda_4*tel_loss + fbcl_loss

            loss.backward()
            optimizers[0].step()
            if batch_idx % args.log_interval == 0:
                log_str = 'Train Audio Model Epoch: {} [{}/{} ({:.0f}%)]\tps_loss: {:.3f}\tmel_loss: {:.3f}\ttel_loss: {:.3f}\tfbcl_loss: {:.3f}'.format(epoch, batch_idx * len(audio), len(train_loader.dataset), 100. * batch_idx / len(train_loader), ps_loss.item(), mel_loss.item(), tel_loss.item(), fbcl_loss.item())
                exp_logger(log_str)
        else: 
            ## Train video model
            global_logits, a_logits, v_logits, frame_logits, temporal_logits, frame_att, temporal_att, sims_after, mask_after = models[1](audio, video, video_st, with_ca=True, epoch_num = epoch-1)

            a_frame_logits = frame_logits[:,:,0,:,:]
            v_frame_logits = frame_logits[:,:,1,:,:]

            fbcl_loss = args.lamda_1*criterion2(sims_after, mask_after) + args.lamda_2*(criterion3(temporal_att[..., 0], frame_att[..., 0].max(dim=-2)[0]) + criterion3(temporal_att[..., 1], frame_att[..., 1].max(dim=-2)[0]))

            frame_target = audio_pseudo_labels * visual_pseudo_labels * target.unsqueeze(dim=1).expand(Pa.shape[0], 10, 25)
            global_sl_loss = criterion_edl(global_logits, target)
            temporal_sl_loss = criterion_edl(temporal_logits, frame_target) 

            tel_loss = global_sl_loss + temporal_sl_loss

            a_frame_sl_loss = criterion_edl(a_frame_logits, audio_pseudo_labels, fore_only = True)
            v_frame_sl_loss = criterion_edl(v_frame_logits, visual_pseudo_labels, fore_only = True)


            ps_loss = args.audio_factor * a_frame_sl_loss + v_frame_sl_loss

            audio_sl_loss = criterion_edl(a_logits, Pa)
            video_sl_loss = criterion_edl(v_logits, Pv)

            mel_loss = audio_sl_loss + args.video_factor * video_sl_loss

            loss = ps_loss + args.lamda_3*mel_loss + args.lamda_4*tel_loss + fbcl_loss

            loss.backward()
            optimizers[1].step()
            if batch_idx % args.log_interval == 0:
                log_str = 'Train Video Model Epoch: {} [{}/{} ({:.0f}%)]\tps_loss: {:.3f}\tmel_loss: {:.3f}\ttel_loss: {:.3f}\tfbcl_loss: {:.3f}'.format(epoch, batch_idx * len(audio), len(train_loader.dataset), 100. * batch_idx / len(train_loader), ps_loss.item(), mel_loss.item(), tel_loss.item(), fbcl_loss.item())
                exp_logger(log_str)


def eval(args, model, val_loader, set, epoch = 0, dual_eval=False):
    if isinstance(model, list):
        for m in model:
            m.eval()
    else:
        model.eval()
    exp_logger("begin evaluate.")
    # load annotations
    df = pd.read_csv(set, header=0, sep='\t')
    df_a = pd.read_csv("data/AVVP_eval_audio.csv", header=0, sep='\t')
    df_v = pd.read_csv("data/AVVP_eval_visual.csv", header=0, sep='\t')

    id_to_idx = {id: index for index, id in enumerate(categories)}
    F_seg_a = []
    F_seg_v = []
    F_seg = []
    F_seg_av = []
    F_event_a = []
    F_event_v = []
    F_event = []
    F_event_av = []

    with torch.no_grad():

        criterion_edl_modality = EvidenceLoss(num_classes=2, evidence = 'exp')
        for batch_idx, sample in tqdm(enumerate(val_loader), total=len(val_loader), desc="Eval epoch {}".format(epoch)):
            audio, video, video_st, target = sample['audio'].to('cuda'), \
                                             sample['video_s'].to('cuda'), \
                                             sample['video_st'].to('cuda'), \
                                             sample['label'].to('cuda')
            
            if isinstance(model, list):
                _, a_logits, _, a_frame_logits, _, _, _, _, _ = model[0](audio, video, video_st)
                _, _, v_logits, v_frame_logits, _, _, _, _, _ = model[1](audio, video, video_st)
                a_sl_prob, _ = criterion_edl_modality.get_predictions(a_logits)
                v_sl_prob, _ = criterion_edl_modality.get_predictions(v_logits)
                a_frame_sl_prob, _ = criterion_edl_modality.get_predictions(a_frame_logits)
                v_frame_sl_prob, _ = criterion_edl_modality.get_predictions(v_frame_logits)
                frame_sl_prob = torch.stack((a_frame_sl_prob[:,:,0,:,:], v_frame_sl_prob[:,:,1,:,:]), dim=2)

            else:
                _, a_logits, v_logits, frame_logits, _, _, _, _, _ = model(audio, video, video_st)
                a_sl_prob, _ = criterion_edl_modality.get_predictions(a_logits)
                v_sl_prob, _ = criterion_edl_modality.get_predictions(v_logits)
                frame_sl_prob, _ = criterion_edl_modality.get_predictions(frame_logits)

            a_prob = torch.clamp(a_sl_prob[:,:,0], min=1e-7, max=1 - 1e-7)
            v_prob = torch.clamp(v_sl_prob[:,:,0], min=1e-7, max=1 - 1e-7)
            a_temporal_prob = frame_sl_prob[:,:,0,:, 0]
            v_temporal_prob = frame_sl_prob[:,:,1,:, 0]


            Pa = a_temporal_prob.cpu().detach().numpy()[0, :, :]
            Pv = v_temporal_prob.cpu().detach().numpy()[0, :, :]
            a_prob = a_prob.cpu().detach().numpy()
            v_prob = v_prob.cpu().detach().numpy()
            a_sl_prob = a_sl_prob.cpu().detach().numpy()
            v_sl_prob = v_sl_prob.cpu().detach().numpy()


            v_thresh_list = np.full((1, 25), 0.45)
            a_thresh_list = np.full((1, 25), 0.45)

            if args.selected_thre:
                excel_path = osp.join(args.model_save_dir, args.checkpoint, args.checkpoint) + '_thresh_a.xlsx'
                wb = load_workbook(excel_path)
                ws = wb['Thres']
                for idx in range(25):
                    a_thresh_list[0, idx] = ws['A'][idx].value

                excel_path = osp.join(args.model_save_dir, args.checkpoint, args.checkpoint) + '_thresh_v.xlsx'
                wb = load_workbook(excel_path)
                ws = wb['Thres']
                for idx in range(25):
                    v_thresh_list[0, idx] = ws['A'][idx].value

            oa = (a_prob >= a_thresh_list).astype(np.int_)
            ov = (v_prob >= v_thresh_list).astype(np.int_)


            # filter out false positive events with predicted weak labels
            Pa = (Pa >= 0.25).astype(np.int_) * np.repeat(oa, repeats=10, axis=0)
            Pv = (Pv >= 0.25).astype(np.int_) * np.repeat(ov, repeats=10, axis=0)

            # extract audio GT labels
            GT_a = np.zeros((25, 10))
            GT_v = np.zeros((25, 10))


            df_vid_a = df_a.loc[df_a['filename'] == df.loc[batch_idx, :][0]]
            filenames = df_vid_a["filename"]
            events = df_vid_a["event_labels"]
            onsets = df_vid_a["onset"]
            offsets = df_vid_a["offset"]
            num = len(filenames)
            if num > 0:
                for i in range(num):
                    x1 = int(onsets[df_vid_a.index[i]])
                    x2 = int(offsets[df_vid_a.index[i]])
                    event = events[df_vid_a.index[i]]
                    idx = id_to_idx[event]
                    GT_a[idx, x1:x2] = 1

            # extract visual GT labels
            df_vid_v = df_v.loc[df_v['filename'] == df.loc[batch_idx, :][0]]
            filenames = df_vid_v["filename"]
            events = df_vid_v["event_labels"]
            onsets = df_vid_v["onset"]
            offsets = df_vid_v["offset"]
            num = len(filenames)
            if num > 0:
                for i in range(num):
                    x1 = int(onsets[df_vid_v.index[i]])
                    x2 = int(offsets[df_vid_v.index[i]])
                    event = events[df_vid_v.index[i]]
                    idx = id_to_idx[event]
                    GT_v[idx, x1:x2] = 1
    
            GT_av = GT_a * GT_v

            # obtain prediction matrices
            SO_a = np.transpose(Pa)
            SO_v = np.transpose(Pv)
            SO_av = SO_a * SO_v

            # segment-level F1 scores
            f_a, f_v, f, f_av = segment_level(SO_a, SO_v, SO_av, GT_a, GT_v, GT_av)
            F_seg_a.append(f_a)
            F_seg_v.append(f_v)
            F_seg.append(f)
            F_seg_av.append(f_av)

            # event-level F1 scores
            f_a, f_v, f, f_av = event_level(SO_a, SO_v, SO_av, GT_a, GT_v, GT_av)
            F_event_a.append(f_a)
            F_event_v.append(f_v)
            F_event.append(f)
            F_event_av.append(f_av)



    audio_segment_level, visual_segment_level, av_segment_level, avg_type, avg_event, \
        audio_event_level, visual_event_level, av_event_level, avg_type_event, avg_event_level \
        = print_overall_metric(F_seg_a, F_seg_v, F_seg, F_seg_av, F_event_a, F_event_v, F_event, F_event_av)
    return audio_segment_level, visual_segment_level, av_segment_level, avg_type, avg_event, \
        audio_event_level, visual_event_level, av_event_level, avg_type_event, avg_event_level


def main():
    # Training settings
    parser = argparse.ArgumentParser(description='PyTorch Implementation of Audio-Visual Video Parsing')
    parser.add_argument("--audio_dir", type=str, default='data/feats/vggish/', help="audio dir")
    parser.add_argument("--video_dir", type=str, default='data/feats/res152/', help="video dir")
    parser.add_argument("--st_dir", type=str, default='data/feats/r2plus1d_18/', help="video dir")
    parser.add_argument("--label_train", type=str, default="data/AVVP_train.csv", help="weak train csv file")
    parser.add_argument("--label_val", type=str, default="data/AVVP_val_pd.csv", help="weak val csv file")
    parser.add_argument("--label_test", type=str, default="data/AVVP_test_pd.csv", help="weak test csv file")
    parser.add_argument('--batch_size', type=int, default=128, help='batch size for training')
    parser.add_argument('--epochs', type=int, default=10, help='number of epochs to train')
    parser.add_argument('--warm_up_epoch', type=float, default=0.9, help='warm-up epochs')
    parser.add_argument('--optimizer', type=str, choices=['sgd', 'adam'], default='adam')
    parser.add_argument('--lr', type=float, default=5e-4, help='learning rate')
    parser.add_argument('--weight_decay', type=float, default=0)
    parser.add_argument('--lr_step_size', type=int, default=6)
    parser.add_argument('--lr_gamma', type=float, default=0.25)
    parser.add_argument('--seed', type=int, default=6, help='random seed')
    parser.add_argument('--eval_interval', type=int, default=1)
    parser.add_argument('--start_class', type=int, default=0)
    parser.add_argument("--mode", type=str, default='train_with_label_smoothing',
                        choices=['train_with_label_smoothing', 'train_with_label_refinement', 'train_with_pseudo_labels', 'test_model', 'select_thresholds'],
                        help="with mode to use")
    parser.add_argument('--num_layers', type=int, default=1)
    parser.add_argument('--noise_ratio_file', type=str, default='noise_ratios.npz')
    parser.add_argument('--a_smooth', type=float, default=1.0)
    parser.add_argument('--v_smooth', type=float, default=0.9)
    parser.add_argument('--clamp', type=float, default=1e-7)
    parser.add_argument('--nce_smooth', type=float, default=0.1)
    parser.add_argument('--max_thresh', type=float, default=0.8)
    parser.add_argument('--min_thresh', type=float, default=0.2)
    parser.add_argument('--temperature', type=float, default=0.2, help='feature temperature number')
    parser.add_argument('--warm_num', type=int, default=1000, help='warm epoch number')
    parser.add_argument('--log_interval', type=int, default=700, help='how many batches for logging training status')
    parser.add_argument('--log_file', type=str, help="log file path")
    parser.add_argument('--save_model', type=str, default="true", choices=["true", "false"], help='whether to save model')
    parser.add_argument("--model_save_dir", type=str, default='models/', help="model save dir")
    parser.add_argument("--checkpoint", type=str, default='NREP_PL', help="save model name")

    parser.add_argument("--logger", type=str, default='logger/', help="save model name")
    parser.add_argument("--best_audio_metric", type=str, default='audio_seg', choices=["audio_seg", "visual_seg", "av_seg", "avg_type_seg", "avg_event_seg", "audio_eve", "visual_eve", "av_eve", "avg_type_eve", "avg_event_eve"], help="best audio metric type")
    parser.add_argument("--best_video_metric", type=str, default='avg_event_eve', choices=["audio_seg", "visual_seg", "av_seg", "avg_type_seg", "avg_event_seg", "audio_eve", "visual_eve", "av_eve", "avg_type_eve", "avg_event_eve"], help="best metric type")
    parser.add_argument("--no-log", action='store_true', default=False, help="logger switcher")
    parser.add_argument("--selected_thre", action='store_true', default=False, help="logger switcher")
    parser.add_argument("--vggish", action='store_true', default=False, help="feature switcher")


    parser.add_argument("--v_pseudo_data_dir", type=str, default='/data/CLIP/segment_pseudo_labels', help="visual segment-level pseudo labels dir")
    parser.add_argument("--a_pseudo_data_dir", type=str, default='/data/CLAP/segment_pseudo_labels', help="audio segment-level pseudo labels dir")

    parser.add_argument('--lamda_1', type=float, default=1.0)
    parser.add_argument('--lamda_2', type=float, default=1.0)
    parser.add_argument('--lamda_3', type=float, default=1.0)
    parser.add_argument('--lamda_4', type=float, default=1.0)
    parser.add_argument('--audio_factor', type=float, default=2.0)
    parser.add_argument('--video_factor', type=float, default=3.0)
    parser.add_argument("--evidence", type=str, default='exp', choices=["relu", "exp", "softplus"])
    parser.add_argument("--loss_type", type=str, default='bce', choices=["bce", "log"])
    parser.add_argument("--uncertain", action='store_true', default=False, help="feature switcher")



    args = parser.parse_args()

    if args.no_log:
        exp_logger.disable_file()
    else:
        os.makedirs(os.path.dirname(args.logger), exist_ok=True)
        logger_timestamps = datetime.datetime.now()
        exp_logger.set_file(os.path.join(args.logger, args.mode+'_'+datetime.datetime.strftime(logger_timestamps,'%Y-%m-%d %H:%M:%S')+".log"))

    save_dir = osp.join(args.model_save_dir, args.checkpoint)
    os.makedirs(save_dir, exist_ok=True)

    # print parameters
    exp_logger('----------------args-----------------')
    for k in list(vars(args).keys()):
        exp_logger('%s: %s' % (k, vars(args)[k]))
    exp_logger('----------------args-----------------')
    cur = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
    exp_logger(f'current time: {cur}')

    set_random_seed(args.seed)
    if args.log_file:
        os.makedirs(os.path.dirname(args.log_file), exist_ok=True)
    save_model = args.save_model == 'true'
    os.makedirs(args.model_save_dir, exist_ok=True)

    start = time.time()

    if args.mode == 'train_with_pseudo_labels':
        logger = SummaryWriter(args.log_file) if args.log_file else None

        args.with_ca = False
        train_loader, val_loader, test_loader = get_LLP_dataloader(args)

        if args.vggish:
            audio_model = NREP(args.num_layers, args.temperature, dim_a=128, dim_v=2048).to('cuda')
            video_model = copy.deepcopy(audio_model)
        else:
            audio_model = NREP(args.num_layers, args.temperature).to('cuda')
            video_model = copy.deepcopy(audio_model)
        models = [audio_model, video_model]


        optimizers = [optim.Adam(audio_model.parameters(), lr=args.lr), optim.Adam(video_model.parameters(), lr=args.lr, weight_decay=args.weight_decay)]
        schedulers = [optim.lr_scheduler.StepLR(optimizers[0], step_size=args.lr_step_size, gamma=args.lr_gamma), optim.lr_scheduler.StepLR(optimizers[1], step_size=args.lr_step_size, gamma=args.lr_gamma)]


        criterion = nn.BCELoss()

        best_audio_F = 0
        best_audio_model = None
        best_video_F = 0
        best_video_model = None
        best_state_dict = get_random_state()
        
        for i, scheduler in enumerate(schedulers):  
            for epoch in range(1, args.epochs + 1):
                train_with_pseudo_labels(args, models, train_loader, optimizers, criterion, epoch=args.epochs*i+epoch, logger=logger)
                scheduler.step(epoch)


                for idx, (model, optimizer) in enumerate(zip(models, optimizers)):
                    
                    state_dict = get_random_state()
                    state_dict['model_{}'.format(idx)] = model.state_dict()
                    state_dict['optimizer_{}'.format(idx)] = optimizer.state_dict()
                    state_dict['scheduler_{}'.format(idx)] = scheduler.state_dict()
                    state_dict['epochs_{}'.format(idx)] = args.epochs

                os.makedirs(os.path.dirname(osp.join(args.model_save_dir, args.checkpoint)), exist_ok=True)
                save_path = osp.join(args.model_save_dir, args.checkpoint, args.checkpoint+'.ckpt')
                exp_logger("Saving checkpoint at epoch {}: {} \n".format(args.epochs*i+epoch, save_path))
                torch.save(state_dict, save_path)

                exp_logger("Validation Performance of Epoch {}:".format(args.epochs*i+epoch))

                audio_seg, visual_seg, av_seg, avg_type_seg, avg_event_seg, audio_eve, visual_eve, av_eve, avg_type_eve, avg_event_eve = eval(args, models, val_loader, args.label_val, args.epochs*i+epoch, dual_eval=True)

                if locals()[args.best_audio_metric] >= best_audio_F:
                    best_audio_F = locals()[args.best_audio_metric]
                    best_audio_model = copy.deepcopy(models[0])
                    if save_model:
                        best_state_dict['model_audio'] = best_audio_model.state_dict()
                        save_path = osp.join(args.model_save_dir, args.checkpoint, args.checkpoint+'.best')
                        torch.save(best_state_dict, save_path)
                        exp_logger("Best audio results at {} have been updated: {} \n".format(args.best_audio_metric, save_path))

                if locals()[args.best_video_metric] >= best_video_F:
                    best_video_F = locals()[args.best_video_metric]
                    best_video_model = copy.deepcopy(models[1])
                    if save_model:
                        best_state_dict['model_video'] = best_video_model.state_dict()
                        save_path = osp.join(args.model_save_dir, args.checkpoint, args.checkpoint+'.best')
                        torch.save(best_state_dict, save_path)
                        exp_logger("Best video results at {} have been updated: {} \n".format(args.best_video_metric, save_path))

        if logger:
            logger.close()
        optimizer.zero_grad()
        models = [best_audio_model, best_video_model]
        exp_logger("Test the best model:")
        eval(args, models, test_loader, args.label_test)

    elif args.mode == 'test_model':
        dataset = args.label_test
        args.with_ca = True 

        audio_model = NREP(args.num_layers, args.temperature).to('cuda')
        video_model = copy.deepcopy(audio_model)

        test_dataset = LLP_dataset(label=args.label_test, audio_dir=args.audio_dir,
                               video_dir=args.video_dir, st_dir=args.st_dir,
                               transform=transforms.Compose([ToTensor()]), v_pseudo_data_dir=args.v_pseudo_data_dir, a_pseudo_data_dir=args.a_pseudo_data_dir)
        test_loader = DataLoader(test_dataset, batch_size=1, shuffle=False, num_workers=1, pin_memory=True)
        
        resume_path = osp.join(args.model_save_dir, args.checkpoint, args.checkpoint+'.best')
        exp_logger("Resuming model from {}".format(resume_path))
        resume = torch.load(resume_path)
        audio_model.load_state_dict(resume['model_audio'])
        video_model.load_state_dict(resume['model_video'])

        models = [audio_model, video_model]
        eval(args, models, test_loader, dataset)


    end = time.time()
    exp_logger(f'duration time {(end - start) / 60} mins.')
    cur = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
    exp_logger(f'current time: {cur}')


if __name__ == '__main__':
    main()
